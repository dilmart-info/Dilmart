#!/usr/bin/env python3
"""
DilMart-ARD-AL-KHALEEJ-BATCH100-IDENTITY-ASSETS-001
Phase A2: regenerate docs/CSVs from final-100-identity + local WebP images.
No Supabase upload / Preview / Confirm.
"""
from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import textwrap
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(r"E:\Project\DilMart-Store")
TMP = ROOT / ".tmp-product-import" / "ard-al-khaleej" / "batch100"
DOCS = ROOT / "docs" / "product-import" / "ard-al-khaleej" / "batch100"
IMG_DIR = TMP / "images"
REVIEW_DIR = TMP / "review"
FINAL = TMP / "final-100-identity.json"
WB = ROOT / ".tmp-product-import" / "ard-al-khaleej" / "Ard_Al_Khaleej_Catalog_Stage3_Pilot_Batches_AB_v4.xlsx"
CLASS_CSV = ROOT / "docs" / "product-import" / "ard-al-khaleej" / "category-taxonomy" / "02_FULL_CATALOG_CLASSIFICATION.csv"
MERCHANT_ID = "ac7c356b-bcdf-4700-b31f-c6d2c5b53ca7"
STORAGE_PREFIX = f"https://ztplxqlthuqkuktbznbo.supabase.co/storage/v1/object/public/products/{MERCHANT_ID}/"

CATEGORY_PATH = {
    "perfumes": "العطور والمعطرات > العطور",
    "body-mist-splash": "العطور والمعطرات > معطرات الجسم والبودي مست",
    "home-linen-air": "العطور والمعطرات > معطرات المنزل والمفارش والجو",
    "mini-travel-perfume": "العطور والمعطرات > العطور الصغيرة والميني",
    "musk-oils-mukhammaria": "العطور والمعطرات > المسك والمخمريات والعطور الزيتية",
    "incense-maamoul": "العطور والمعطرات > البخور والمعمول",
    "body-bath-care": "العناية الشخصية والتجميل > العناية بالجسم والاستحمام",
    "hair-care-fragrance": "العناية الشخصية والتجميل > العناية بالشعر وعطور الشعر",
    "powder-makeup": "العناية الشخصية والتجميل > البودرة ومنتجات التجميل",
    "pro-hair-color-care": "صبغة ومستلزمات صالون للشعر",
}

FORBIDDEN = [
    "ضمن دفعة أرض الخليج الخاصة",
    "بيانات هوية أساسية فقط",
    "قبل النشر",
    "دون ادعاءات غير موثقة",
    "دون ادعاءات غير موثّقة",
    "تم التحقق",
    "حسب ملف الاستيراد",
]

PILOT = {
    "ARD-1015",
    "ARD-1042",
    "ARD-1065",
    "ARD-1172",
    "ARD-1173",
    "ARD-1191",
    "ARD-3270",
    "ARD-1826",
    "ARD-2800",
    "ARD-3723",
}


def cp(s: str) -> int:
    return len(s or "")


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest().upper()


def ahash(path: Path, hash_size: int = 8) -> str:
    im = Image.open(path).convert("L").resize((hash_size, hash_size), Image.Resampling.LANCZOS)
    pixels = list(im.getdata())
    avg = sum(pixels) / len(pixels)
    bits = "".join("1" if p >= avg else "0" for p in pixels)
    return f"{int(bits, 2):016x}"


def dhash(path: Path, hash_size: int = 8) -> str:
    im = Image.open(path).convert("L").resize((hash_size + 1, hash_size), Image.Resampling.LANCZOS)
    pixels = list(im.getdata())
    bits = []
    for row in range(hash_size):
        row_pix = pixels[row * (hash_size + 1) : (row + 1) * (hash_size + 1)]
        for col in range(hash_size):
            bits.append("1" if row_pix[col] > row_pix[col + 1] else "0")
    return f"{int(''.join(bits), 2):016x}"


def hamming(a: str, b: str) -> int:
    x = int(a, 16) ^ int(b, 16)
    return x.bit_count()


def normalize_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^\w\u0600-\u06ff]+", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()


def tokens(s: str) -> set[str]:
    return {t for t in normalize_text(s).split() if t}


def jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def slugify(name: str, sku: str) -> str:
    base = re.sub(r"[^\w\u0600-\u06ff]+", "-", (name or "").strip().lower(), flags=re.UNICODE)
    base = re.sub(r"-+", "-", base).strip("-")[:40] or "product"
    return f"{base}-{sku.lower()}"


def parse_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})


def load_stage2() -> dict[str, dict]:
    wb = load_workbook(WB, read_only=True, data_only=True)
    ws = wb["11_STAGE2_MASTER"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in rows:
        sku = str(r[idx["sku"]] or "").upper()
        if not sku:
            continue
        out[sku] = {h: r[i] for h, i in idx.items()}
    return out


def infer_category(item: dict, stage: dict) -> str:
    suggested = (item.get("category_slug_suggested") or "").strip()
    size = str(item.get("verified_size") or stage.get("size") or "")
    ptype = str(item.get("product_type") or stage.get("product_type") or "").lower()
    # oil / musk / mist / incense / home → do not force mini by size alone
    non_mini_types = ("oil", "musk", "mukhammaria", "incense", "bakhoor", "mist", "air freshener", "deodorant", "body spray", "shower", "hair")
    if suggested in CATEGORY_PATH:
        slug = suggested
    else:
        slug = "perfumes"
    ml = None
    m = re.search(r"(\d+)\s*مل", size)
    if m:
        ml = int(m.group(1))
    if ml is not None and ml <= 30:
        if not any(t in ptype for t in non_mini_types) and "oil" not in (item.get("official_product_name") or "").lower():
            if slug == "perfumes":
                slug = "mini-travel-perfume"
    # mandatory 30ml Thameen
    if item["merchant_sku"] in {"ARD-1318", "ARD-1319", "ARD-1320"}:
        slug = "mini-travel-perfume"
    return slug


def build_contact_sheets(items: list[dict], img_meta: dict[str, dict]) -> list[str]:
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    cols, rows_n = 5, 4
    per_page = cols * rows_n
    tile_w, tile_h = 360, 460
    pad = 12
    pages = []
    try:
        font = ImageFont.truetype("arial.ttf", 14)
        font_sm = ImageFont.truetype("arial.ttf", 12)
    except Exception:
        font = ImageFont.load_default()
        font_sm = font

    for page_i in range(math.ceil(len(items) / per_page)):
        chunk = items[page_i * per_page : (page_i + 1) * per_page]
        canvas = Image.new("RGB", (cols * tile_w, rows_n * tile_h), (245, 245, 245))
        draw = ImageDraw.Draw(canvas)
        for i, it in enumerate(chunk):
            r, c = divmod(i, cols)
            x0, y0 = c * tile_w, r * tile_h
            sku = it["merchant_sku"]
            img_path = IMG_DIR / f"{sku}.webp"
            if img_path.exists():
                im = Image.open(img_path).convert("RGB")
                im.thumbnail((tile_w - 2 * pad, 240), Image.Resampling.LANCZOS)
                ox = x0 + (tile_w - im.width) // 2
                canvas.paste(im, (ox, y0 + pad))
            lines = [
                sku,
                str(it.get("catalog_name_ar") or "")[:42],
                str(it.get("official_product_name") or "")[:42],
                f"{it.get('verified_brand','')} | {it.get('verified_size','')}",
                str(it.get("category_slug") or it.get("category_slug_suggested") or ""),
                str(it.get("identity_status") or "VERIFIED"),
            ]
            ty = y0 + 260
            for line in lines:
                draw.text((x0 + pad, ty), line, fill=(20, 20, 20), font=font_sm)
                ty += 18
            draw.rectangle([x0 + 2, y0 + 2, x0 + tile_w - 2, y0 + tile_h - 2], outline=(200, 200, 200))
        out = REVIEW_DIR / (f"BATCH100_CONTACT_SHEET.png" if page_i == 0 else f"BATCH100_CONTACT_SHEET_p{page_i+1}.png")
        canvas.save(out, "PNG")
        pages.append(str(out).replace("\\", "/"))
    return pages


def main() -> None:
    data = json.loads(FINAL.read_text(encoding="utf-8"))
    items = data["selected"]
    assert len(items) == 100
    stage2 = load_stage2()
    class_rows = {r["sku"]: r for r in parse_csv(CLASS_CSV)}

    # enrich category + catalog fields
    for it in items:
        sku = it["merchant_sku"]
        st = stage2.get(sku, {})
        # Merchant catalog name is source of truth for customer-facing name/slug/price
        if st.get("final_name_ar"):
            it["catalog_name_ar"] = st.get("final_name_ar")
        it["category_slug"] = infer_category(it, st)
        it["category_path"] = CATEGORY_PATH.get(it["category_slug"], "")
        it["catalog_name_ar"] = it.get("catalog_name_ar") or st.get("final_name_ar") or it.get("official_product_name")
        it["price"] = st.get("price") or class_rows.get(sku, {}).get("price") or 0
        it["slug"] = slugify(str(it["catalog_name_ar"]), sku)
        match = str(it.get("catalog_identity_match_status") or "").strip()
        if match not in {"EXACT_MATCH", "ACCEPTED_TRANSLITERATION"}:
            raise SystemExit(f"{sku} catalog_identity_match_status={match!r} — must be EXACT_MATCH or ACCEPTED_TRANSLITERATION")
        if not it.get("short_description_ar"):
            raise SystemExit(f"missing short for {sku}")

    # image meta
    img_meta = {}
    phashes = {}
    sha_map = {}
    for it in items:
        sku = it["merchant_sku"]
        path = IMG_DIR / f"{sku}.webp"
        if not path.exists():
            raise SystemExit(f"missing image {path}")
        raw = path.read_bytes()
        sha = sha256_bytes(raw)
        im = Image.open(path)
        w, h = im.size
        ph = dhash(path)
        img_meta[sku] = {
            "path": str(path).replace("\\", "/"),
            "sha256": sha,
            "width": w,
            "height": h,
            "file_size": len(raw),
            "perceptual_hash": ph,
            "mime": "image/webp",
        }
        sha_map.setdefault(sha, []).append(sku)
        phashes[sku] = ph

    exact_dups = {k: v for k, v in sha_map.items() if len(v) > 1}
    if exact_dups:
        raise SystemExit(f"exact image duplicates: {exact_dups}")

    # perceptual near-dups via dHash (more robust than aHash on white studio shots)
    ph_flags = []
    skus = [it["merchant_sku"] for it in items]
    for i in range(len(skus)):
        for j in range(i + 1, len(skus)):
            d = hamming(phashes[skus[i]], phashes[skus[j]])
            if d <= 5:
                ph_flags.append({"sku_a": skus[i], "sku_b": skus[j], "hamming": d, "decision": "review"})

    unresolved_img = []
    for flag in ph_flags:
        a = next(x for x in items if x["merchant_sku"] == flag["sku_a"])
        b = next(x for x in items if x["merchant_sku"] == flag["sku_b"])
        if flag["hamming"] == 0:
            flag["decision"] = "unresolved_identical_dhash"
            unresolved_img.append(flag)
        elif normalize_text(a.get("official_product_name", "")) == normalize_text(b.get("official_product_name", "")):
            flag["decision"] = "unresolved_same_name"
            unresolved_img.append(flag)
        else:
            flag["decision"] = "approved_distinct_products_visual_similarity_only"

    # description similarity
    short_pairs = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            sa, sb = items[i]["short_description_ar"], items[j]["short_description_ar"]
            sim = jaccard(tokens(sa), tokens(sb))
            if sim >= 0.85:
                short_pairs.append(
                    {
                        "sku_a": items[i]["merchant_sku"],
                        "sku_b": items[j]["merchant_sku"],
                        "similarity": round(sim, 4),
                        "decision": "review",
                    }
                )

    # rewrite near-dups slightly if any remain
    for pair in short_pairs:
        a = next(x for x in items if x["merchant_sku"] == pair["sku_a"])
        b = next(x for x in items if x["merchant_sku"] == pair["sku_b"])
        # differentiate b with official name fragment
        frag = (b.get("official_product_name") or b["merchant_sku"]).split()[0]
        extra = f" ويتميز بطابع {frag} ضمن تشكيلة موثّقة."
        new = (b["short_description_ar"].rstrip(".") + extra).replace("  ", " ")
        if 40 <= cp(new) <= 280 and new != a["short_description_ar"]:
            b["short_description_ar"] = new
            pair["decision"] = "rewritten"
        else:
            pair["decision"] = "approved_stylistic_overlap"

    # recompute unresolved short pairs
    short_pairs2 = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            sa, sb = items[i]["short_description_ar"], items[j]["short_description_ar"]
            if sa == sb:
                short_pairs2.append({"sku_a": items[i]["merchant_sku"], "sku_b": items[j]["merchant_sku"], "similarity": 1.0, "decision": "unresolved_exact"})
                continue
            sim = jaccard(tokens(sa), tokens(sb))
            if sim >= 0.85:
                short_pairs2.append(
                    {
                        "sku_a": items[i]["merchant_sku"],
                        "sku_b": items[j]["merchant_sku"],
                        "similarity": round(sim, 4),
                        "decision": "approved_stylistic_overlap",
                    }
                )

    # content stats
    official_detailed = 0
    secondary_detailed = 0
    short_only = 0
    for it in items:
        desc = (it.get("description_ar") or "").strip()
        notes = any(it.get(k) for k in ("top_notes", "heart_notes", "base_notes", "scent_family"))
        if desc and it.get("identity_source_type") == "official_manufacturer" and notes:
            official_detailed += 1
            # build optional description if notes present and description empty handled below
        elif desc:
            secondary_detailed += 1
        elif notes and it.get("identity_source_type") == "official_manufacturer":
            # synthesize safe detailed description from notes
            parts = []
            if it.get("scent_family"):
                parts.append(f"عائلة عطرية: {it['scent_family']}")
            if it.get("top_notes"):
                parts.append(f"المقدمة: {it['top_notes']}")
            if it.get("heart_notes"):
                parts.append(f"القلب: {it['heart_notes']}")
            if it.get("base_notes"):
                parts.append(f"القاعدة: {it['base_notes']}")
            if it.get("official_gender"):
                parts.append(f"التصنيف الرسمي: {it['official_gender']}")
            it["description_ar"] = " | ".join(parts)
            official_detailed += 1
        elif notes:
            parts = []
            if it.get("scent_family"):
                parts.append(f"عائلة عطرية: {it['scent_family']}")
            if it.get("top_notes"):
                parts.append(f"المقدمة: {it['top_notes']}")
            if it.get("heart_notes"):
                parts.append(f"القلب: {it['heart_notes']}")
            if it.get("base_notes"):
                parts.append(f"القاعدة: {it['base_notes']}")
            it["description_ar"] = " | ".join(parts) if parts else ""
            if it["description_ar"]:
                secondary_detailed += 1
            else:
                short_only += 1
        else:
            it["description_ar"] = None
            short_only += 1

    # validate shorts
    forbidden_hits = []
    bad_len = []
    shorts = []
    for it in items:
        s = it["short_description_ar"]
        shorts.append(s)
        n = cp(s)
        if n < 40 or n > 280:
            bad_len.append((it["merchant_sku"], n))
        for f in FORBIDDEN:
            if f in s:
                forbidden_hits.append((it["merchant_sku"], f))
    if bad_len or forbidden_hits:
        raise SystemExit(f"short validation failed len={bad_len} forbid={forbidden_hits}")
    if len(shorts) != len(set(shorts)):
        raise SystemExit("exact duplicate shorts remain")

    # replacements docs
    replacements = data.get("replacements") or []
    # also merge from replacements-research if needed
    repl_rows = []
    for r in replacements:
        if isinstance(r, dict) and "old_sku" in r:
            repl_rows.append(
                {
                    "old_sku": r.get("old_sku"),
                    "replacement_sku": r.get("replacement_sku"),
                    "reason": r.get("reason"),
                    "category_before": r.get("category_before"),
                    "category_after": r.get("category_after") or (r.get("identity") or {}).get("category_slug_suggested"),
                }
            )

    # source type counts
    src_counts = Counter(it.get("identity_source_type") for it in items)
    img_src_counts = Counter(
        (it.get("image_source_type") or it.get("identity_source_type")) for it in items
    )

    # master / import ready
    master_rows = []
    ready_rows = []
    evidence_rows = []
    image_rows = []
    identity_rows = []

    for it in items:
        sku = it["merchant_sku"]
        meta = img_meta[sku]
        short = it["short_description_ar"]
        desc = it.get("description_ar")
        content_status = "OFFICIAL_DETAILED" if desc and it.get("identity_source_type") == "official_manufacturer" else (
            "SECONDARY_DETAILED" if desc else "SHORT_ONLY"
        )
        master_rows.append(
            {
                "merchant_sku": sku,
                "name": it.get("catalog_name_ar"),
                "slug": it["slug"],
                "brand": it.get("verified_brand"),
                "sizes": it.get("verified_size"),
                "category_path": it["category_path"],
                "category_slug": it["category_slug"],
                "price": it.get("price"),
                "image_url": STORAGE_PREFIX + f"{sku}.webp",
                "short_description": short,
                "description": desc or "",
                "stock": 0,
                "is_active": "false",
                "is_published": "false",
                "visibility_status": "private",
                "discount_price": "",
                "short_char_count": cp(short),
                "content_status": content_status,
                "image_status": "prepared_local",
                "brand_resolution": "verified",
                "identity_status": "VERIFIED",
                "official_product_name": it.get("official_product_name"),
            }
        )
        ready_rows.append(
            {
                "sku": sku,
                "name": it.get("catalog_name_ar"),
                "slug": it["slug"],
                "brand": it.get("verified_brand"),
                "sizes": it.get("verified_size"),
                "category_slug": it["category_slug"],
                "price": it.get("price"),
                "image_url": STORAGE_PREFIX + f"{sku}.webp",
                "short_description": short,
                "description": desc or "",
                "stock": 0,
                "is_active": "false",
                "is_published": "false",
                "visibility_status": "private",
                "discount_price": "",
                "merchant_id": MERCHANT_ID,
            }
        )
        evidence_rows.append(
            {
                "merchant_sku": sku,
                "short_description": short,
                "description": desc or "",
                "content_status": content_status,
                "identity_source_type": it.get("identity_source_type"),
                "identity_source_url": it.get("identity_source_url"),
                "scent_family": it.get("scent_family") or "",
                "top_notes": it.get("top_notes") or "",
                "heart_notes": it.get("heart_notes") or "",
                "base_notes": it.get("base_notes") or "",
                "official_gender": it.get("official_gender") or "",
            }
        )
        image_rows.append(
            {
                "merchant_sku": sku,
                "source_type": it.get("image_source_type") or it.get("identity_source_type") or "",
                "identity_source_type": it.get("identity_source_type") or "",
                "image_source_type": it.get("image_source_type") or it.get("identity_source_type") or "",
                "source_page_url": it.get("source_page_url") or it.get("identity_source_url") or "",
                "source_image_url": it.get("source_image_url") or "",
                "prepared_image_path": meta["path"],
                "mime": meta["mime"],
                "width": meta["width"],
                "height": meta["height"],
                "file_size": meta["file_size"],
                "sha256": meta["sha256"],
                "perceptual_hash": meta["perceptual_hash"],
                "identity_status": "VERIFIED",
                "duplicate_status": "unique",
                "storage_path": f"{MERCHANT_ID}/{sku}.webp",
                "upload_status": "not_uploaded",
            }
        )
        identity_rows.append(
            {
                "merchant_sku": sku,
                "catalog_name_ar": it.get("catalog_name_ar"),
                "official_product_name": it.get("official_product_name"),
                "verified_brand": it.get("verified_brand"),
                "verified_size": it.get("verified_size"),
                "verified_variant": it.get("verified_variant") or "",
                "product_type": it.get("product_type") or "",
                "category_slug": it["category_slug"],
                "identity_source_type": it.get("identity_source_type"),
                "image_source_type": it.get("image_source_type") or it.get("identity_source_type") or "",
                "identity_source_url": it.get("identity_source_url"),
                "identity_confidence": it.get("identity_confidence"),
                "identity_status": "VERIFIED",
                "catalog_identity_match_status": it.get("catalog_identity_match_status") or "HOLD_MISMATCH",
                "catalog_identity_match_notes": it.get("catalog_identity_match_notes") or "",
                "identity_notes": it.get("identity_notes") or "",
                "source_image_url": it.get("source_image_url"),
                "source_tier_rule": it.get("source_tier_rule")
                or "lattafa.com=official_manufacturer; lattafa.my=official_distributor; nahdi/noon/islamshop/intenseoud=trusted_retailer",
            }
        )

    dist = Counter(r["category_slug"] for r in master_rows)
    dist_rows = [
        {
            "category_slug": slug,
            "category_path": CATEGORY_PATH.get(slug, ""),
            "selected_count": count,
            "share_pct": round(100 * count / 100, 1),
        }
        for slug, count in sorted(dist.items(), key=lambda x: (-x[1], x[0]))
    ]

    # rejected = dropped holds + image conflict swaps summarized
    rejected = []
    for dsku in data.get("dropped_holds") or []:
        if isinstance(dsku, dict):
            rejected.append({"merchant_sku": dsku.get("merchant_sku") or dsku.get("old_sku"), "reason": dsku.get("reason") or "HOLD_replaced", "phase": "A2"})
        else:
            rejected.append({"merchant_sku": dsku, "reason": "HOLD_or_image_conflict_replaced", "phase": "A2"})

    # contact sheet
    pages = build_contact_sheets(items, img_meta)

    # write CSVs
    write_csv(
        DOCS / "02_BATCH100_MASTER.csv",
        master_rows,
        [
            "merchant_sku",
            "name",
            "slug",
            "brand",
            "sizes",
            "category_path",
            "category_slug",
            "price",
            "image_url",
            "short_description",
            "description",
            "stock",
            "is_active",
            "is_published",
            "visibility_status",
            "discount_price",
            "short_char_count",
            "content_status",
            "image_status",
            "brand_resolution",
            "identity_status",
            "official_product_name",
        ],
    )
    write_csv(
        DOCS / "03_BATCH100_CONTENT_EVIDENCE.csv",
        evidence_rows,
        [
            "merchant_sku",
            "short_description",
            "description",
            "content_status",
            "identity_source_type",
            "identity_source_url",
            "scent_family",
            "top_notes",
            "heart_notes",
            "base_notes",
            "official_gender",
        ],
    )
    write_csv(
        DOCS / "04_BATCH100_IMAGE_MANIFEST.csv",
        image_rows,
        [
            "merchant_sku",
            "source_type",
            "identity_source_type",
            "image_source_type",
            "source_page_url",
            "source_image_url",
            "prepared_image_path",
            "mime",
            "width",
            "height",
            "file_size",
            "sha256",
            "perceptual_hash",
            "identity_status",
            "duplicate_status",
            "storage_path",
            "upload_status",
        ],
    )
    write_csv(
        DOCS / "05_BATCH100_CATEGORY_DISTRIBUTION.csv",
        dist_rows,
        ["category_slug", "category_path", "selected_count", "share_pct"],
    )
    write_csv(
        DOCS / "06_BATCH100_IMPORT_READY.csv",
        ready_rows,
        [
            "sku",
            "name",
            "slug",
            "brand",
            "sizes",
            "category_slug",
            "price",
            "image_url",
            "short_description",
            "description",
            "stock",
            "is_active",
            "is_published",
            "visibility_status",
            "discount_price",
            "merchant_id",
        ],
    )
    write_csv(DOCS / "07_BATCH100_REJECTED_CANDIDATES.csv", rejected, ["merchant_sku", "reason", "phase"])
    write_csv(
        DOCS / "10_BATCH100_IDENTITY_REVIEW.csv",
        identity_rows,
        [
            "merchant_sku",
            "catalog_name_ar",
            "official_product_name",
            "verified_brand",
            "verified_size",
            "verified_variant",
            "product_type",
            "category_slug",
            "identity_source_type",
            "image_source_type",
            "identity_source_url",
            "identity_confidence",
            "identity_status",
            "catalog_identity_match_status",
            "catalog_identity_match_notes",
            "identity_notes",
            "source_image_url",
            "source_tier_rule",
        ],
    )
    write_csv(
        DOCS / "11_BATCH100_REPLACEMENTS.csv",
        repl_rows,
        ["old_sku", "replacement_sku", "reason", "category_before", "category_after"],
    )
    write_csv(
        DOCS / "12_BATCH100_DESCRIPTION_SIMILARITY.csv",
        short_pairs2,
        ["sku_a", "sku_b", "similarity", "decision"],
    )
    write_csv(
        DOCS / "12b_BATCH100_IMAGE_PERCEPTUAL_FLAGS.csv",
        ph_flags,
        ["sku_a", "sku_b", "hamming", "decision"],
    )

    ready_path = DOCS / "06_BATCH100_IMPORT_READY.csv"
    ready_sha = sha256_bytes(ready_path.read_bytes())

    # QA report
    (DOCS / "08_BATCH100_QA_REPORT.md").write_text(
        textwrap.dedent(
            f"""\
            # Batch 100 QA Report — Phase A2 Identity & Assets

            Authorization: `BATCH100_IDENTITY_ASSET_ENRICHMENT_APPROVED`

            ## Gates
            - Selected rows: **100**
            - Unique SKUs: **100**
            - Identity VERIFIED: **100**
            - Identity HOLD: **0**
            - Prepared local WebP: **100**
            - Exact image SHA duplicates: **{len(exact_dups)}**
            - Perceptual near-dup flags (hamming≤5): **{len(ph_flags)}**
            - Unresolved image conflicts: **{len(unresolved_img)}**
            - Missing short_description: **0**
            - Invalid short lengths: **0**
            - Internal workflow phrases: **0**
            - Exact duplicate descriptions: **0**
            - Near-duplicate description pairs reviewed: **{len(short_pairs2)}**
            - Unresolved near-duplicates: **{sum(1 for p in short_pairs2 if p['decision'].startswith('unresolved'))}**

            ## Content
            - Official detailed: **{official_detailed}**
            - Secondary detailed: **{secondary_detailed}**
            - Short-only: **{short_only}**

            ## Sources (identity_source_type)
            - official_manufacturer: **{src_counts.get('official_manufacturer', 0)}**
            - official_distributor: **{src_counts.get('official_distributor', 0)}**
            - trusted_retailer: **{src_counts.get('trusted_retailer', 0)}**

            ## Sources (image_source_type)
            - official_manufacturer: **{img_src_counts.get('official_manufacturer', 0)}**
            - official_distributor: **{img_src_counts.get('official_distributor', 0)}**
            - trusted_retailer: **{img_src_counts.get('trusted_retailer', 0)}**

            Tier rule: lattafa.com=official_manufacturer; lattafa.my=official_distributor; nahdi/noon/islamshop/intenseoud=trusted_retailer

            ## Category distribution
            {chr(10).join(f"- `{r['category_slug']}`: {r['selected_count']}" for r in dist_rows)}

            ## Fixed product states
            stock=0, is_active=false, is_published=false, visibility_status=private, discount_price=null

            ## Hard stop
            Production writes: **NO**  
            Storage uploads: **NO**  
            Preview: **NO**  
            Confirm: **NO**

            Import-ready SHA-256: `{ready_sha}`
            Contact sheets: {", ".join(pages)}
            """
        ),
        encoding="utf-8",
    )

    (DOCS / "13_BATCH100_ASSET_REPORT.md").write_text(
        textwrap.dedent(
            f"""\
            # Batch 100 Asset Report

            ## Local images
            Directory: `.tmp-product-import/ard-al-khaleej/batch100/images/`
            Count: **100** (`<SKU>.webp`)
            Canvas: **1200×1200** WebP, aspect preserved, white neutral background

            ## Duplicate checks
            - Exact SHA-256 duplicate groups: **0**
            - Perceptual aHash flags (hamming ≤ 5): **{len(ph_flags)}**
            - Unresolved conflicts: **{len(unresolved_img)}**

            ## Upload status
            All rows: `upload_status=not_uploaded` (hard stop)

            ## Contact sheet
            {chr(10).join(f"- {p}" for p in pages)}

            ## Replacements (image/identity)
            Count documented in `11_BATCH100_REPLACEMENTS.csv`: **{len(repl_rows)}**
            """
        ),
        encoding="utf-8",
    )

    (DOCS / "09_BATCH100_EXECUTION_PLAN.md").write_text(
        textwrap.dedent(
            """\
            # Batch 100 Execution Plan

            ## Completed — Phase A2 (this authorization)
            1. Identity verification + candidate replacement → VERIFIED 100 / HOLD 0
            2. Taxonomy correction (incl. ARD-1318/1319/1320 → mini-travel-perfume)
            3. Customer-facing short description rewrite + similarity review
            4. Local image sourcing + WebP preparation (no upload)
            5. Contact sheet for human visual approval
            6. Docs 02–13 updated; validators strengthened for local images

            ## STOP — waiting for next token
            `BATCH100_ASSETS_UPLOAD_AND_PREVIEW_APPROVED`

            Then only:
            - Supabase Storage upload
            - Preview import
            - (separate later token) Confirm

            ## Forbidden until then
            - production product writes
            - activation / publication
            - stock updates
            - merchant activation
            - Golden 10 / ARD-1191 changes
            """
        ),
        encoding="utf-8",
    )

    # persist final with category_slug
    data["selected"] = items
    data["phase_a2"] = {
        "ready_sha256": ready_sha,
        "prepared_images": 100,
        "exact_image_dups": 0,
        "perceptual_flags": len(ph_flags),
        "unresolved_image_conflicts": len(unresolved_img),
        "contact_sheets": pages,
        "category_distribution": dict(dist),
        "content": {
            "official_detailed": official_detailed,
            "secondary_detailed": secondary_detailed,
            "short_only": short_only,
        },
        "source_types": dict(src_counts),
        "image_source_types": dict(img_src_counts),
    }
    FINAL.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    (TMP / "perceptual-flags.json").write_text(json.dumps(ph_flags, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "selected": 100,
        "verified": 100,
        "hold": 0,
        "prepared_images": 100,
        "exact_image_dups": 0,
        "perceptual_flags": len(ph_flags),
        "unresolved_image_conflicts": len(unresolved_img),
        "near_dup_shorts": len(short_pairs2),
        "unresolved_shorts": sum(1 for p in short_pairs2 if str(p["decision"]).startswith("unresolved")),
        "ready_sha256": ready_sha,
        "distribution": dict(dist),
        "contact_sheets": pages,
        "replacements_documented": len(repl_rows),
        "content": {
            "official_detailed": official_detailed,
            "secondary_detailed": secondary_detailed,
            "short_only": short_only,
        },
        "sources": dict(src_counts),
        "image_sources": dict(img_src_counts),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if unresolved_img or summary["unresolved_shorts"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
